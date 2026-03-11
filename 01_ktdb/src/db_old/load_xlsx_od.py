"""
KTDB OD 데이터 PostgreSQL 적재 스크립트
- Task 1: 비수도권 목적OD → ktdb_od_purpose_subzone (5 regions × 7 years)
- Task 2: 전국 250존 목적OD → ktdb_od_purpose_250 (7 years)
- Task 3: 화물 OD → ktdb_od_freight (1 base + 6 future years)

psycopg v3 COPY protocol 사용 (bulk insert)
openpyxl read_only=True, data_only=True
"""

from __future__ import annotations

import io
import os
import sys
import time
from pathlib import Path
from typing import Iterator

import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Project root & env
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")

import psycopg  # noqa: E402  (after dotenv)


# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------

def get_conn() -> psycopg.Connection:
    conninfo = (
        f"host={os.environ['DB_HOST']} "
        f"port={os.environ.get('DB_PORT', '5432')} "
        f"dbname={os.environ['DB_NAME']} "
        f"user={os.environ['DB_USER']} "
        f"password={os.environ['DB_PASSWORD']} "
        f"sslmode=require"
    )
    return psycopg.connect(conninfo, prepare_threshold=None)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _float(v) -> float | None:
    """Convert cell value to float, None if empty/invalid."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def elapsed(start: float) -> str:
    s = time.time() - start
    return f"{s:.1f}s"


# ---------------------------------------------------------------------------
# TASK 1: 비수도권 목적OD → ktdb_od_purpose_subzone
# ---------------------------------------------------------------------------

SUBZONE_FILES = {
    "02": PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-02 부산울산권 목적 OD(2023-2050)/2024-OD-PSN-OBJ-02 부산울산권 목적 OD(2023-2050)/4. OD/부산울산권_목적통행.xlsx",
    "03": PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-03 대구광역권 목적 OD(2023-2050)/2024-OD-PSN-OBJ-03 대구광역권 목적 OD(2023-2050)/4. OD/대구광역권_목적통행.xlsx",
    "04": PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-04 대전세종충청권목적 OD(2023-2050)/2024-OD-PSN-OBJ-04 대전세종충청권목적 OD(2023-2050)/4. OD/대전세종충청권_목적통행.xlsx",
    "05": PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-05 광주광역권 목적 OD(2023-2050)/2024-OD-PSN-OBJ-05 광주광역권 목적 OD(2023-2050)/4. OD/광주광역권_목적통행.xlsx",
    "06": PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-06 제주권 목적 OD(2023-2050)/2024-OD-PSN-OBJ-06 제주권 목적 OD(2023-2050)/4. OD/제주권_목적통행.xlsx",
}

YEAR_SHEETS = ["2023년", "2025년", "2030년", "2035년", "2040년", "2045년", "2050년"]


def _already_loaded_subzone(cur: psycopg.Cursor, year: int, region_id: str) -> bool:
    cur.execute(
        "SELECT 1 FROM ktdb_od_purpose_subzone WHERE year=%s AND region_id=%s LIMIT 1",
        (year, region_id),
    )
    return cur.fetchone() is not None


def _rows_subzone(
    ws, year: int, region_id: str, source_file: str
) -> Iterator[tuple]:
    """Yield COPY rows from one year-sheet of a subzone file."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        seq_o = _int(row[0])
        seq_d = _int(row[1])
        if seq_o is None or seq_d is None:
            continue
        origin_zone = f"{region_id}_{seq_o:04d}"
        dest_zone = f"{region_id}_{seq_d:04d}"
        commute = _float(row[2]) or 0.0
        school = _float(row[3]) or 0.0
        business = _float(row[4]) or 0.0
        return_home = _float(row[5]) or 0.0
        etc = _float(row[6]) or 0.0
        total = _float(row[7]) or 0.0
        yield (year, region_id, origin_zone, dest_zone,
               commute, school, business, return_home, etc, total, source_file)


def load_subzone_od(conn: psycopg.Connection) -> int:
    """Load Task 1: 비수도권 목적OD."""
    print("\n=== TASK 1: 비수도권 목적OD → ktdb_od_purpose_subzone ===")
    total_rows = 0
    t0 = time.time()

    with conn.cursor() as cur:
        for region_id, xlsx_path in SUBZONE_FILES.items():
            region_names = {"02": "부산울산권", "03": "대구광역권", "04": "대전세종충청권",
                            "05": "광주광역권", "06": "제주권"}
            region_name = region_names[region_id]
            print(f"\n  [{region_id}] {region_name}: {xlsx_path.name}")

            if not xlsx_path.exists():
                print(f"    ERROR: 파일 없음 → {xlsx_path}")
                continue

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            for sheet_name in YEAR_SHEETS:
                # Find matching sheet by year number (handles encoding issues)
                year = int(sheet_name.replace("년", ""))
                # Try exact match first, fallback to index-based match
                ws = None
                if sheet_name in sheet_names:
                    ws = wb[sheet_name]
                else:
                    # Match by position: sheets[1..7] = 2023,2025,2030,2035,2040,2045,2050
                    year_idx = YEAR_SHEETS.index(sheet_name)
                    if year_idx + 1 < len(sheet_names):
                        ws = wb[sheet_names[year_idx + 1]]
                        print(f"    Using sheet index {year_idx+1} for {sheet_name}")

                if ws is None:
                    print(f"    SKIP: sheet '{sheet_name}' not found")
                    continue

                if _already_loaded_subzone(cur, year, region_id):
                    print(f"    SKIP: year={year} region={region_id} already loaded")
                    continue

                t1 = time.time()
                rows = list(_rows_subzone(ws, year, region_id, xlsx_path.name))
                n = len(rows)

                if n == 0:
                    print(f"    SKIP: no data rows for year={year}")
                    continue

                # COPY bulk insert
                with cur.copy(
                    "COPY ktdb_od_purpose_subzone "
                    "(year, region_id, origin_zone, destination_zone, "
                    "commute, school, business, return_home, etc, total, source_file) "
                    "FROM STDIN"
                ) as copy:
                    for row in rows:
                        copy.write_row(row)

                conn.commit()
                total_rows += n
                print(f"    year={year}: {n:,} rows in {elapsed(t1)}")

            wb.close()

    print(f"\n  TASK 1 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# TASK 2: 전국 250존 목적OD → ktdb_od_purpose_250
# ---------------------------------------------------------------------------

OD_250_FILE = PROJECT_ROOT / "data/extracted/2024-OD-PSN-OBJ-00 전국지역간 목적 OD(250존)(2023-2050)/5. OD/전국_2023년~2050년 목적별OD(250).xlsx"

YEAR_250_SHEETS = [
    ("2023_목적OD", 2023),
    ("2025_목적OD", 2025),
    ("2030_목적OD", 2030),
    ("2035_목적OD", 2035),
    ("2040_목적OD", 2040),
    ("2045_목적OD", 2045),
    ("2050_목적OD", 2050),
]


def _already_loaded_250(cur: psycopg.Cursor, year: int) -> bool:
    cur.execute(
        "SELECT 1 FROM ktdb_od_purpose_250 WHERE year=%s LIMIT 1",
        (year,),
    )
    return cur.fetchone() is not None


def _rows_250(ws, year: int, source_file: str) -> Iterator[tuple]:
    """Yield COPY rows from one year-sheet of the 250존 file."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        origin_sido = _int(row[0])
        dest_sido = _int(row[2])
        origin_sigungu = _int(row[1])  # 250존 code for origin
        dest_sigungu = _int(row[3])    # 250존 code for dest
        # Note: column order is 출발시도 | 출발시군구 | 도착시도 | 도착시군구
        # but task spec says: 출발시도 | 도착시도 | 출발시군구 | 도착시군구
        # Actual data row confirmed: (1,1,1,1,...) - let's match header order from inspection
        # Row 1 header: 출발시도 | 도착시도 | 출발시군구 | 도착시군구 | 출근 | ...
        # So: col0=출발시도, col1=도착시도, col2=출발시군구, col3=도착시군구
        origin_sido = _int(row[0])
        dest_sido = _int(row[1])
        origin_sigungu = _int(row[2])
        dest_sigungu = _int(row[3])
        if origin_sido is None or dest_sido is None:
            continue
        commute = _float(row[4]) or 0.0
        school = _float(row[5]) or 0.0
        business = _float(row[6]) or 0.0
        return_home = _float(row[7]) or 0.0
        etc = _float(row[8]) or 0.0
        total = _float(row[9]) or 0.0
        yield (year, str(origin_sido), str(origin_sigungu),
               str(dest_sido), str(dest_sigungu),
               commute, school, business, return_home, etc, total, source_file)


def load_od_250(conn: psycopg.Connection) -> int:
    """Load Task 2: 전국 250존 목적OD."""
    print("\n=== TASK 2: 전국 250존 목적OD → ktdb_od_purpose_250 ===")
    total_rows = 0
    t0 = time.time()

    if not OD_250_FILE.exists():
        print(f"  ERROR: 파일 없음 → {OD_250_FILE}")
        return 0

    print(f"  파일: {OD_250_FILE.name}")

    with conn.cursor() as cur:
        wb = openpyxl.load_workbook(OD_250_FILE, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"  시트 목록: {sheet_names}")

        for sheet_name, year in YEAR_250_SHEETS:
            # Find matching sheet (encoding-safe by index)
            ws = None
            if sheet_name in sheet_names:
                ws = wb[sheet_name]
            else:
                # Sheets are ordered; 목적OD sheets are the last 7
                purp_sheets = [s for s in sheet_names if "OD" in s or "od" in s.lower()]
                idx = [2023, 2025, 2030, 2035, 2040, 2045, 2050].index(year)
                if idx < len(purp_sheets):
                    ws = wb[purp_sheets[idx]]
                    print(f"  Using sheet '{purp_sheets[idx]}' for year={year}")

            if ws is None:
                # Fallback: last 7 sheets are the 목적OD sheets
                all_sheets = sheet_names
                if len(all_sheets) >= 7:
                    idx = [2023, 2025, 2030, 2035, 2040, 2045, 2050].index(year)
                    offset = len(all_sheets) - 7
                    ws = wb[all_sheets[offset + idx]]
                    print(f"  Fallback sheet '{all_sheets[offset + idx]}' for year={year}")

            if ws is None:
                print(f"  SKIP: sheet for year={year} not found")
                continue

            if _already_loaded_250(cur, year):
                print(f"  SKIP: year={year} already loaded")
                continue

            t1 = time.time()
            rows = list(_rows_250(ws, year, OD_250_FILE.name))
            n = len(rows)

            if n == 0:
                print(f"  SKIP: no data rows for year={year}")
                continue

            with cur.copy(
                "COPY ktdb_od_purpose_250 "
                "(year, origin_sido, origin_sigungu, destination_sido, destination_sigungu, "
                "commute, school, business, return_home, etc, total, source_file) "
                "FROM STDIN"
            ) as copy:
                for row in rows:
                    copy.write_row(row)

            conn.commit()
            total_rows += n
            print(f"  year={year}: {n:,} rows in {elapsed(t1)}")

        wb.close()

    print(f"\n  TASK 2 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# TASK 3: 화물 OD → ktdb_od_freight
# ---------------------------------------------------------------------------

FREIGHT_BASE_FILE = PROJECT_ROOT / "data/extracted/2024-OD-FRE-CAR-00 전국지역간 톤급별 화물자동차 통행량 OD(2023년 기준)/배포용 (기준년도 2023년) 화물자동차OD_2025.11.28.xlsx"
FREIGHT_FUTURE_FILE = PROJECT_ROOT / "data/extracted/2024-OD-FRE-CAR-00 전국지역간 톤급별 화물자동차 통행량 OD(2023년 기준)/배포용 장래년도 화물자동차OD_2025.11.28.xlsx"

FREIGHT_FUTURE_YEARS = [2025, 2030, 2035, 2040, 2045, 2050]


def _already_loaded_freight(cur: psycopg.Cursor, year: int) -> bool:
    cur.execute(
        "SELECT 1 FROM ktdb_od_freight WHERE year=%s LIMIT 1",
        (year,),
    )
    return cur.fetchone() is not None


def _rows_freight(ws, year: int, source_file: str, skip_rows: int = 2) -> Iterator[tuple]:
    """
    Yield COPY rows from one freight sheet.
    skip_rows: 1 = title only (future sheets already skipped), 2 = title + header
    Row layout after skipping: O_250 | 대존O_17 | D_250 | 대존D_17 | 소형 | 중형 | 대형 | 전체
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        o250 = _int(row[0])
        o17 = _int(row[1])
        d250 = _int(row[2])
        d17 = _int(row[3])
        if o250 is None or d250 is None:
            continue
        small = _float(row[4]) or 0.0
        medium = _float(row[5]) or 0.0
        large = _float(row[6]) or 0.0
        total = _float(row[7]) or 0.0
        yield (year, o250, o17, d250, d17, small, medium, large, total, source_file)


def load_freight_od(conn: psycopg.Connection) -> int:
    """Load Task 3: 화물 OD."""
    print("\n=== TASK 3: 화물 OD → ktdb_od_freight ===")
    total_rows = 0
    t0 = time.time()

    COPY_SQL = (
        "COPY ktdb_od_freight "
        "(year, origin_zone_250, origin_zone_17, destination_zone_250, destination_zone_17, "
        "tonnage_small, tonnage_medium, tonnage_large, tonnage_total, source_file) "
        "FROM STDIN"
    )

    with conn.cursor() as cur:
        # --- Base year 2023 ---
        if not FREIGHT_BASE_FILE.exists():
            print(f"  ERROR: 기준년도 파일 없음 → {FREIGHT_BASE_FILE}")
        else:
            print(f"\n  기준년도 파일: {FREIGHT_BASE_FILE.name}")
            if _already_loaded_freight(cur, 2023):
                print("  SKIP: year=2023 already loaded")
            else:
                wb = openpyxl.load_workbook(FREIGHT_BASE_FILE, read_only=True, data_only=True)
                # First sheet is the data sheet (2023년)
                ws = wb.worksheets[0]
                t1 = time.time()
                # Row 1 = title, Row 2 = header → skip 2 rows
                rows = list(_rows_freight(ws, 2023, FREIGHT_BASE_FILE.name, skip_rows=2))
                n = len(rows)
                if n > 0:
                    with cur.copy(COPY_SQL) as copy:
                        for row in rows:
                            copy.write_row(row)
                    conn.commit()
                    total_rows += n
                    print(f"  year=2023: {n:,} rows in {elapsed(t1)}")
                else:
                    print("  SKIP: no data rows for year=2023")
                wb.close()

        # --- Future years ---
        if not FREIGHT_FUTURE_FILE.exists():
            print(f"  ERROR: 장래년도 파일 없음 → {FREIGHT_FUTURE_FILE}")
        else:
            print(f"\n  장래년도 파일: {FREIGHT_FUTURE_FILE.name}")
            wb = openpyxl.load_workbook(FREIGHT_FUTURE_FILE, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            print(f"  시트 목록: {sheet_names}")

            for year in FREIGHT_FUTURE_YEARS:
                year_str = f"{year}년"
                ws = None
                if year_str in sheet_names:
                    ws = wb[year_str]
                else:
                    # Match by index among first 6 sheets
                    idx = FREIGHT_FUTURE_YEARS.index(year)
                    if idx < len(sheet_names):
                        ws = wb[sheet_names[idx]]
                        print(f"  Using sheet '{sheet_names[idx]}' for year={year}")

                if ws is None:
                    print(f"  SKIP: sheet for year={year} not found")
                    continue

                if _already_loaded_freight(cur, year):
                    print(f"  SKIP: year={year} already loaded")
                    continue

                t1 = time.time()
                # Row 1 = title, Row 2 = header → skip 2
                rows = list(_rows_freight(ws, year, FREIGHT_FUTURE_FILE.name, skip_rows=2))
                n = len(rows)
                if n == 0:
                    print(f"  SKIP: no data rows for year={year}")
                    continue

                with cur.copy(COPY_SQL) as copy:
                    for row in rows:
                        copy.write_row(row)

                conn.commit()
                total_rows += n
                print(f"  year={year}: {n:,} rows in {elapsed(t1)}")

            wb.close()

    print(f"\n  TASK 3 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("KTDB OD 데이터 PostgreSQL 적재 시작")
    print(f"PROJECT_ROOT: {PROJECT_ROOT}")
    t_start = time.time()

    conn = get_conn()
    print("DB 연결 성공")

    try:
        r1 = load_subzone_od(conn)
        r2 = load_od_250(conn)
        r3 = load_freight_od(conn)
    finally:
        conn.close()

    print("\n" + "=" * 60)
    print("전체 완료")
    print(f"  Task 1 (비수도권 목적OD):  {r1:>10,} rows")
    print(f"  Task 2 (전국 250존 목적OD): {r2:>10,} rows")
    print(f"  Task 3 (화물 OD):           {r3:>10,} rows")
    print(f"  합계:                        {r1+r2+r3:>10,} rows")
    print(f"  총 소요시간: {elapsed(t_start)}")


if __name__ == "__main__":
    main()
