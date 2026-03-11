"""
KTDB 수단별(mode) OD 데이터 PostgreSQL 적재 스크립트

PART 1: 전국 250존 수단OD → ktdb_od_mode_250
PART 2: 비수도권 수단OD (xlsx) → ktdb_od_mode_subzone (long format)
PART 3: 수도권 수단OD (TXT, CP949) → ktdb_od_mode_subzone (long format)

psycopg v3 COPY protocol 사용 (bulk insert)
openpyxl read_only=True, data_only=True
"""

from __future__ import annotations

import os
import sys
import time
from pathlib import Path
from typing import Iterator

import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Project root & env
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")

import psycopg  # noqa: E402  (after dotenv)


# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------

def get_conn() -> psycopg.Connection:
    conninfo = (
        f"host={os.environ['DB_HOST']} "
        f"port={os.environ.get('DB_PORT', '5432')} "
        f"dbname={os.environ['DB_NAME']} "
        f"user={os.environ['DB_USER']} "
        f"password={os.environ['DB_PASSWORD']} "
        f"sslmode=require"
    )
    return psycopg.connect(conninfo, prepare_threshold=None)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def elapsed(start: float) -> str:
    s = time.time() - start
    return f"{s:.1f}s"


# ---------------------------------------------------------------------------
# PART 1: 전국 250존 수단OD → ktdb_od_mode_250
# ---------------------------------------------------------------------------

OD_MODE_250_FILE = PROJECT_ROOT / (
    "data/extracted/"
    "2024-OD-PSN-MOD-10 전국지역간 주수단 OD(250존)(2023-2050)/"
    "5. OD/"
    "전국_2023년~2050년 주수단별OD(250).xlsx"
)

# Sheet names: '2023_주수단OD' through '2050_주수단OD'
YEAR_250_MODE_SHEETS = [
    ("2023_주수단OD", 2023),
    ("2025_주수단OD", 2025),
    ("2030_주수단OD", 2030),
    ("2035_주수단OD", 2035),
    ("2040_주수단OD", 2040),
    ("2045_주수단OD", 2045),
    ("2050_주수단OD", 2050),
]

COPY_SQL_250 = (
    "COPY ktdb_od_mode_250 "
    "(year, origin_sido, origin_sigungu, destination_sido, destination_sigungu, "
    "car, bus, subway, rail, ktx, air, ship, total, source_file) "
    "FROM STDIN"
)


def _delete_mode_250(cur: psycopg.Cursor, year: int) -> int:
    cur.execute("DELETE FROM ktdb_od_mode_250 WHERE year=%s", (year,))
    return cur.rowcount


def _rows_mode_250(ws, year: int, source_file: str) -> Iterator[tuple]:
    """
    Yield rows from one year-sheet of the 250존 mode file.
    Columns (0-indexed):
      0=출발시도, 1=도착시도, 2=출발시군구, 3=도착시군구,
      4=승용차, 5=버스, 6=지하철, 7=일반철도, 8=고속철도, 9=항공, 10=해운, 11=합계
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        origin_sido = _int(row[0])
        dest_sido = _int(row[1])
        origin_sigungu = _int(row[2])
        dest_sigungu = _int(row[3])
        if origin_sido is None or dest_sido is None:
            continue
        car = _float(row[4]) or 0.0
        bus = _float(row[5]) or 0.0
        subway = _float(row[6]) or 0.0
        rail = _float(row[7]) or 0.0
        ktx = _float(row[8]) or 0.0
        air = _float(row[9]) or 0.0
        ship = _float(row[10]) or 0.0
        total = _float(row[11]) or 0.0
        yield (
            year,
            str(origin_sido), str(origin_sigungu),
            str(dest_sido), str(dest_sigungu),
            car, bus, subway, rail, ktx, air, ship, total,
            source_file,
        )


def load_mode_od_250(conn: psycopg.Connection) -> int:
    """PART 1: 전국 250존 수단OD."""
    print("\n=== PART 1: 전국 250존 수단OD → ktdb_od_mode_250 ===")
    total_rows = 0
    t0 = time.time()

    if not OD_MODE_250_FILE.exists():
        print(f"  ERROR: 파일 없음 → {OD_MODE_250_FILE}")
        return 0

    print(f"  파일: {OD_MODE_250_FILE.name}")

    with conn.cursor() as cur:
        wb = openpyxl.load_workbook(OD_MODE_250_FILE, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"  시트 목록: {sheet_names}")

        for sheet_name, year in YEAR_250_MODE_SHEETS:
            ws = None
            if sheet_name in sheet_names:
                ws = wb[sheet_name]
            else:
                # Fallback: match by position among data sheets
                idx = [s[1] for s in YEAR_250_MODE_SHEETS].index(year)
                if idx < len(sheet_names):
                    ws = wb[sheet_names[idx]]
                    print(f"  Using sheet '{sheet_names[idx]}' for year={year}")

            if ws is None:
                print(f"  SKIP: sheet for year={year} not found")
                continue

            deleted = _delete_mode_250(cur, year)
            conn.commit()
            if deleted > 0:
                print(f"  Cleared {deleted:,} existing rows for year={year}")

            t1 = time.time()
            rows = list(_rows_mode_250(ws, year, OD_MODE_250_FILE.name))
            n = len(rows)

            if n == 0:
                print(f"  SKIP: no data rows for year={year}")
                continue

            with cur.copy(COPY_SQL_250) as copy:
                for row in rows:
                    copy.write_row(row)

            conn.commit()
            total_rows += n
            print(f"  year={year}: {n:,} rows in {elapsed(t1)}")

        wb.close()

    print(f"\n  PART 1 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# PART 2: 비수도권 수단OD (xlsx) → ktdb_od_mode_subzone (long format)
# ---------------------------------------------------------------------------

SUBZONE_MODE_FILES = {
    "02": PROJECT_ROOT / (
        "data/extracted/"
        "2024-OD-PSN-MOD-12 부산울산권 주수단 OD(2023-2050)/"
        "2024-OD-PSN-MOD-12 부산울산권 주수단 OD(2023-2050)/"
        "4. OD/부산울산권_주수단통행.xlsx"
    ),
    "03": PROJECT_ROOT / (
        "data/extracted/"
        "2024-OD-PSN-MOD-13 대구광역권 주수단 OD(2023-2050)/"
        "2024-OD-PSN-MOD-13 대구광역권 주수단 OD(2023-2050)/"
        "4. OD/대구광역권_주수단통행.xlsx"
    ),
    "04": PROJECT_ROOT / (
        "data/extracted/"
        "2024-OD-PSN-MOD-14 대전세종충청권 주수단 OD(2023-2050)/"
        "2024-OD-PSN-MOD-14 대전세종충청권 주수단 OD(2023-2050)/"
        "4. OD/대전세종충청권_주수단통행.xlsx"
    ),
    "05": PROJECT_ROOT / (
        "data/extracted/"
        "2024-OD-PSN-MOD-15 광주광역권 주수단 OD(2023-2050)/"
        "2024-OD-PSN-MOD-15 광주광역권 주수단 OD(2023-2050)/"
        "4. OD/광주광역권_주수단통행.xlsx"
    ),
    "06": PROJECT_ROOT / (
        "data/extracted/"
        "2024-OD-PSN-MOD-16 제주권 주수단OD(2023-2050)/"
        "2024-OD-PSN-MOD-16 제주권 주수단OD(2023-2050)/"
        "4. OD/제주권_주수단통행.xlsx"
    ),
}

SUBZONE_REGION_NAMES = {
    "02": "부산울산권",
    "03": "대구광역권",
    "04": "대전세종충청권",
    "05": "광주광역권",
    "06": "제주권",
}

# 7 year sheets per file
YEAR_SHEETS_SUBZONE = ["2023년", "2025년", "2030년", "2035년", "2040년", "2045년", "2050년"]

# Mode mapping: (column_index_in_sheet, mode_code)
# Columns: 0=출발, 1=도착, 2=도보/자전거, 3=승용차, 4=버스,
#          5=일반/고속철도, 6=도시철도, 7=택시, 8=기타,
#          9=승용차/택시(skip), 10=기타합계(skip)
SUBZONE_MODES = [
    (2, "walk_bike"),
    (3, "car"),
    (4, "bus"),
    (5, "rail"),
    (6, "metro"),
    (7, "taxi"),
    (8, "etc"),
]

COPY_SQL_MODE_SUBZONE = (
    "COPY ktdb_od_mode_subzone "
    "(year, region_id, origin_zone, destination_zone, mode_code, volume, source_file) "
    "FROM STDIN"
)


def _count_mode_subzone(cur: psycopg.Cursor, year: int, region_id: str) -> int:
    cur.execute(
        "SELECT COUNT(*) FROM ktdb_od_mode_subzone WHERE year=%s AND region_id=%s",
        (year, region_id),
    )
    row = cur.fetchone()
    return row[0] if row else 0


def _delete_mode_subzone(cur: psycopg.Cursor, year: int, region_id: str) -> int:
    cur.execute(
        "DELETE FROM ktdb_od_mode_subzone WHERE year=%s AND region_id=%s",
        (year, region_id),
    )
    return cur.rowcount


def _rows_mode_subzone(
    ws, year: int, region_id: str, source_file: str
) -> Iterator[tuple]:
    """
    Yield long-format rows from one year-sheet of a subzone mode file.
    Each OD pair becomes 7 mode rows (one per mode).
    Skips rows where volume == 0.0.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        seq_o = _int(row[0])
        seq_d = _int(row[1])
        if seq_o is None or seq_d is None:
            continue
        origin_zone = f"{region_id}_{seq_o:04d}"
        dest_zone = f"{region_id}_{seq_d:04d}"
        for col_idx, mode_code in SUBZONE_MODES:
            volume = _float(row[col_idx]) or 0.0
            if volume == 0.0:
                continue
            yield (year, region_id, origin_zone, dest_zone, mode_code, volume, source_file)


def load_mode_subzone_od(conn: psycopg.Connection) -> int:
    """PART 2: 비수도권 수단OD (xlsx)."""
    print("\n=== PART 2: 비수도권 수단OD → ktdb_od_mode_subzone ===")
    total_rows = 0
    t0 = time.time()

    with conn.cursor() as cur:
        for region_id, xlsx_path in SUBZONE_MODE_FILES.items():
            region_name = SUBZONE_REGION_NAMES[region_id]
            print(f"\n  [{region_id}] {region_name}: {xlsx_path.name}")

            if not xlsx_path.exists():
                print(f"    ERROR: 파일 없음 → {xlsx_path}")
                continue

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            for sheet_name in YEAR_SHEETS_SUBZONE:
                year = int(sheet_name.replace("년", ""))
                ws = None
                if sheet_name in sheet_names:
                    ws = wb[sheet_name]
                else:
                    # Fallback: match by position (skip first sheet if it's a summary)
                    year_idx = YEAR_SHEETS_SUBZONE.index(sheet_name)
                    if year_idx + 1 < len(sheet_names):
                        ws = wb[sheet_names[year_idx + 1]]
                        print(f"    Using sheet index {year_idx+1} for {sheet_name}")

                if ws is None:
                    print(f"    SKIP: sheet '{sheet_name}' not found")
                    continue

                # Always delete first (handles partial loads), then commit before COPY
                deleted = _delete_mode_subzone(cur, year, region_id)
                conn.commit()
                if deleted > 0:
                    print(f"    Cleared {deleted:,} existing rows for year={year} region={region_id}")

                t1 = time.time()
                rows = list(_rows_mode_subzone(ws, year, region_id, xlsx_path.name))
                n = len(rows)

                if n == 0:
                    print(f"    SKIP: no data rows for year={year}")
                    continue

                with cur.copy(COPY_SQL_MODE_SUBZONE) as copy:
                    for row in rows:
                        copy.write_row(row)

                conn.commit()
                total_rows += n
                print(f"    year={year}: {n:,} mode rows in {elapsed(t1)}")

            wb.close()

    print(f"\n  PART 2 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# PART 3: 수도권 수단OD (TXT, CP949) → ktdb_od_mode_subzone
# ---------------------------------------------------------------------------

TXT_BASE_DIR = PROJECT_ROOT / (
    "data/extracted/"
    "2024-OD-PSN-MOD-11 수도권 주수단 OD(2023-2050)/"
    "2024-OD-PSN-MOD-11 수도권 주수단 OD(2023-2050)/"
    "4. OD"
)

# Filename suffix → year mapping
TXT_YEAR_MAP = {
    "OD_MMODE_23_F.TXT": 2023,
    "OD_MMODE_25_F.TXT": 2025,
    "OD_MMODE_30_F.TXT": 2030,
    "OD_MMODE_35_F.TXT": 2035,
    "OD_MMODE_40_F.TXT": 2040,
    "OD_MMODE_45_F.TXT": 2045,
    "OD_MMODE_50_F.TXT": 2050,
}

# 10 mode codes for positional mapping (fields 4-13, 0-indexed)
TXT_MODES = [
    "mode_01",  # val1  (승용차/car)
    "mode_02",  # val2  (버스/bus)
    "mode_03",  # val3  (도시철도/metro)
    "mode_04",  # val4
    "mode_05",  # val5
    "mode_06",  # val6  (도보/walk)
    "mode_07",  # val7
    "mode_08",  # val8
    "mode_09",  # val9
    "mode_10",  # val10
]

REGION_ID_METRO = "01"
PROGRESS_INTERVAL = 500_000  # print every N source lines


def _count_metro(cur: psycopg.Cursor, year: int) -> int:
    cur.execute(
        "SELECT COUNT(*) FROM ktdb_od_mode_subzone WHERE year=%s AND region_id=%s",
        (year, REGION_ID_METRO),
    )
    row = cur.fetchone()
    return row[0] if row else 0


def _stream_txt_rows(
    txt_path: Path, year: int, source_file: str
) -> Iterator[tuple[int, list[tuple]]]:
    """
    Stream long-format rows from one CP949 TXT file.
    Yields (line_count, batch_rows) every PROGRESS_INTERVAL lines.
    Fields per line: O_seq O_zone D_seq D_zone val1..val10
    Skips zero-volume rows.
    """
    batch: list[tuple] = []
    line_count = 0

    with open(txt_path, encoding="cp949", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) < 14:
                continue
            line_count += 1

            # O_zone_code and D_zone_code are the zone identifiers
            o_zone = parts[1]
            d_zone = parts[3]

            for i, mode_code in enumerate(TXT_MODES):
                try:
                    volume = float(parts[4 + i])
                except (IndexError, ValueError):
                    continue
                if volume == 0.0:
                    continue
                batch.append((
                    year, REGION_ID_METRO, o_zone, d_zone,
                    mode_code, volume, source_file,
                ))

            if line_count % PROGRESS_INTERVAL == 0:
                yield line_count, batch
                batch = []

    # yield remaining
    yield line_count, batch


def load_mode_metro_od(conn: psycopg.Connection) -> int:
    """PART 3: 수도권 수단OD (TXT)."""
    print("\n=== PART 3: 수도권 수단OD (TXT) → ktdb_od_mode_subzone ===")
    total_rows = 0
    t0 = time.time()

    with conn.cursor() as cur:
        for filename, year in TXT_YEAR_MAP.items():
            txt_path = TXT_BASE_DIR / filename
            print(f"\n  [{year}] {filename}")

            if not txt_path.exists():
                print(f"    ERROR: 파일 없음 → {txt_path}")
                continue

            deleted = _delete_mode_subzone(cur, year, REGION_ID_METRO)
            conn.commit()
            if deleted > 0:
                print(f"    Cleared {deleted:,} existing rows for year={year}")

            t1 = time.time()
            year_rows = 0

            with cur.copy(COPY_SQL_MODE_SUBZONE) as copy:
                for line_count, batch in _stream_txt_rows(txt_path, year, filename):
                    for row in batch:
                        copy.write_row(row)
                    year_rows += len(batch)
                    if line_count % PROGRESS_INTERVAL == 0:
                        print(
                            f"    {line_count:,} source lines processed, "
                            f"{year_rows:,} mode rows so far ... {elapsed(t1)}"
                        )

            conn.commit()
            total_rows += year_rows
            print(f"    year={year}: {year_rows:,} mode rows in {elapsed(t1)}")

    print(f"\n  PART 3 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("KTDB 수단별 OD 데이터 PostgreSQL 적재 시작")
    print(f"PROJECT_ROOT: {PROJECT_ROOT}")
    t_start = time.time()

    conn = get_conn()
    print("DB 연결 성공")

    try:
        r1 = load_mode_od_250(conn)
        r2 = load_mode_subzone_od(conn)
        r3 = load_mode_metro_od(conn)
    finally:
        conn.close()

    print("\n" + "=" * 60)
    print("전체 완료")
    print(f"  PART 1 (전국 250존 수단OD):     {r1:>12,} rows")
    print(f"  PART 2 (비수도권 수단OD xlsx):  {r2:>12,} rows")
    print(f"  PART 3 (수도권 수단OD TXT):     {r3:>12,} rows")
    print(f"  합계:                            {r1+r2+r3:>12,} rows")
    print(f"  총 소요시간: {elapsed(t_start)}")


if __name__ == "__main__":
    main()
