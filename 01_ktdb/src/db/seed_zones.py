"""
존 마스터 데이터 시드 스크립트
- ktdb_zones_17 → ktdb_zones_250 → ktdb_zones_subzone 순서로 적재
- psycopg v3, sslmode='require', ON CONFLICT DO NOTHING (멱등성 보장)
"""

import os
import sys
from pathlib import Path

from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
env_path = PROJECT_ROOT / ".env"
load_dotenv(dotenv_path=env_path)

import openpyxl
import psycopg

# ---------------------------------------------------------------------------
# 데이터 파일 경로
# ---------------------------------------------------------------------------
DATA_ROOT = PROJECT_ROOT / "data" / "extracted"

FILE_250 = (
    DATA_ROOT
    / "2024-OD-PSN-MOD-10 전국지역간 주수단 OD(250존)(2023-2050)"
    / "1. 존체계"
    / "250 존체계.xlsx"
)

FILE_SUDOGWON = (
    DATA_ROOT
    / "2024-OD-PSN-OBJ-01 수도권 목적OD(2023-2050)"
    / "2024-OD-PSN-OBJ-01 수도권 목적OD(2023-2050)"
    / "1. 존체계"
    / "존체계.xlsx"
)

FILE_BISUDO = (
    DATA_ROOT
    / "2024-OD-PSN-OBJ-02 부산울산권 목적 OD(2023-2050)"
    / "2024-OD-PSN-OBJ-02 부산울산권 목적 OD(2023-2050)"
    / "1. 존체계"
    / "존체계.xlsx"
)

# 비수도권 시트명 → region_id 매핑
BISUDO_SHEET_REGION: dict[str, str] = {
    "부산울산권": "02",
    "대구광역권": "03",
    "대전세종충청권": "04",
    "광주광역권": "05",
    "제주권": "06",
}

# 17존 코드 → 시도명 매핑
ZONE17_NAME: dict[int, str] = {
    1: "서울",
    2: "부산",
    3: "대구",
    4: "인천",
    5: "광주",
    6: "대전",
    7: "울산",
    8: "경기",
    9: "강원",
    10: "충북",
    11: "충남",
    12: "전북",
    13: "전남",
    14: "경북",
    15: "경남",
    16: "제주",
    17: "세종",
}


# ---------------------------------------------------------------------------
# DB 연결
# ---------------------------------------------------------------------------

def get_connection() -> psycopg.Connection:
    host = os.environ["DB_HOST"]
    port = os.environ.get("DB_PORT", "5432")
    dbname = os.environ["DB_NAME"]
    user = os.environ["DB_USER"]
    password = os.environ["DB_PASSWORD"]
    conninfo = (
        f"host={host} port={port} dbname={dbname} "
        f"user={user} password={password} sslmode=require connect_timeout=10"
    )
    # prepare_threshold=None disables auto-prepare; needed for PgBouncer
    # in transaction pooling mode where prepared statements cause errors
    return psycopg.connect(conninfo, prepare_threshold=None)


# ---------------------------------------------------------------------------
# 1. ktdb_zones_17
# ---------------------------------------------------------------------------

def seed_zones_17(cur: psycopg.Cursor) -> None:
    print("[1/3] ktdb_zones_17 적재 중...")
    wb = openpyxl.load_workbook(FILE_250, read_only=True, data_only=True)
    ws = wb.active

    seen: set[int] = set()
    params: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] is None:
            continue
        zone_17 = int(row[4])
        if zone_17 in seen:
            continue
        seen.add(zone_17)
        sido_name = ZONE17_NAME.get(zone_17, f"unknown_{zone_17}")
        params.append((zone_17, sido_name))

    wb.close()
    cur.executemany(
        "INSERT INTO ktdb_zones_17 (zone_code_17, sido_name) VALUES (%s, %s) ON CONFLICT (zone_code_17) DO NOTHING",
        params,
    )
    print(f"  -> ktdb_zones_17: {len(params)}행 처리 완료")


# ---------------------------------------------------------------------------
# 2. ktdb_zones_250
# ---------------------------------------------------------------------------

def seed_zones_250(cur: psycopg.Cursor) -> None:
    print("[2/3] ktdb_zones_250 적재 중...")
    wb = openpyxl.load_workbook(FILE_250, read_only=True, data_only=True)
    ws = wb.active

    params: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 컬럼: 대존(시도명) | 소존(시군구명) | 250존 | 161존 | 17존
        if row[2] is None:
            continue
        sido_name: str = str(row[0]).strip() if row[0] else ""
        zone_name: str = str(row[1]).strip() if row[1] else ""
        zone_250: int = int(row[2])
        zone_161 = int(row[3]) if row[3] is not None else None
        zone_17 = int(row[4]) if row[4] is not None else None
        params.append((zone_250, zone_name, sido_name, zone_161, zone_17))

    wb.close()
    cur.executemany(
        "INSERT INTO ktdb_zones_250 (zone_code_250, zone_name, sido_name, zone_161, zone_17) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (zone_code_250) DO NOTHING",
        params,
    )
    print(f"  -> ktdb_zones_250: {len(params)}행 처리 완료")


# ---------------------------------------------------------------------------
# 3. ktdb_zones_subzone
# ---------------------------------------------------------------------------

def _build_sigungu_to_250(cur: psycopg.Cursor) -> dict[str, int]:
    """zone_name(시군구명) → zone_code_250 매핑 딕셔너리 생성."""
    cur.execute("SELECT zone_name, zone_code_250 FROM ktdb_zones_250")
    return {row[0]: row[1] for row in cur.fetchall()}


def seed_subzones_sudogwon(cur: psycopg.Cursor, sigungu_to_250: dict[str, int]) -> int:
    """수도권 소존 적재 (region_id='01')."""
    print("  [수도권] 적재 중...")
    wb = openpyxl.load_workbook(FILE_SUDOGWON, read_only=True, data_only=True)
    ws = wb["존체계 양식"]

    params: list[tuple] = []
    unmatched_sigungu: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 컬럼: 시도 | 시군구 | 행정동 | 권역 존체계_읍면동(seq) | 행정기관코드_읍면동 | 권역내부=1,권역외부=2
        if row[3] is None:
            continue
        sido_name: str = str(row[0]).strip() if row[0] else ""
        sigungu_name: str = str(row[1]).strip() if row[1] else ""
        dong_name: str = str(row[2]).strip() if row[2] else ""
        zone_seq: int = int(row[3])
        admin_code_raw = row[4]
        admin_code: str = str(int(admin_code_raw)) if admin_code_raw is not None else ""

        # zone_code = 행정기관코드 문자열 전체
        zone_code: str = admin_code

        # zone_250 FK: 시군구명으로 매핑
        zone_250 = sigungu_to_250.get(sigungu_name)
        if zone_250 is None:
            unmatched_sigungu.add(sigungu_name)

        params.append((zone_code, zone_seq, dong_name, sigungu_name, sido_name, admin_code, "01", zone_250))

    wb.close()
    if unmatched_sigungu:
        print(f"  [경고] zone_250 매핑 실패 시군구 ({len(unmatched_sigungu)}개): {sorted(unmatched_sigungu)[:10]}")

    cur.executemany(
        "INSERT INTO ktdb_zones_subzone (zone_code, zone_seq, dong_name, sigungu_name, sido_name, admin_code, region_id, zone_250) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (zone_code) DO NOTHING",
        params,
    )
    print(f"  -> 수도권 소존: {len(params)}행 처리 완료")
    return len(params)


def seed_subzones_bisudo(cur: psycopg.Cursor) -> int:
    """비수도권 소존 적재 (region_id='02'~'06')."""
    print("  [비수도권] 적재 중...")
    wb = openpyxl.load_workbook(FILE_BISUDO, read_only=True, data_only=True)

    total_inserted = 0

    for sheet_name, region_id in BISUDO_SHEET_REGION.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [경고] 시트 없음: {sheet_name}")
            continue

        ws = wb[sheet_name]
        params: list[tuple] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 컬럼: 시도 | 시군구 | 행정동 | 권역 존체계_읍면동(seq) | 전국지역간 존체계_시군구(250존) | 권역내부=1,권역외부=2
            if row[3] is None:
                continue
            sido_name: str = str(row[0]).strip() if row[0] else ""
            sigungu_name: str = str(row[1]).strip() if row[1] else ""
            dong_name: str = str(row[2]).strip() if row[2] else ""
            zone_seq: int = int(row[3])
            zone_250_raw = row[4]
            zone_250: int | None = int(zone_250_raw) if zone_250_raw is not None else None

            # zone_code: region_id + seq (4자리 zero-padded)
            zone_code: str = f"{region_id}_{zone_seq:04d}"
            params.append((zone_code, zone_seq, dong_name, sigungu_name, sido_name, region_id, zone_250))

        cur.executemany(
            "INSERT INTO ktdb_zones_subzone (zone_code, zone_seq, dong_name, sigungu_name, sido_name, region_id, zone_250) VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT (zone_code) DO NOTHING",
            params,
        )
        print(f"  -> {sheet_name} (region_id={region_id}): {len(params)}행 처리 완료")
        total_inserted += len(params)

    wb.close()
    return total_inserted


def seed_zones_subzone(cur: psycopg.Cursor) -> None:
    print("[3/3] ktdb_zones_subzone 적재 중...")
    sigungu_to_250 = _build_sigungu_to_250(cur)
    print(f"  ktdb_zones_250에서 {len(sigungu_to_250)}개 시군구 매핑 로드 완료")

    n_sudo = seed_subzones_sudogwon(cur, sigungu_to_250)
    n_bisudo = seed_subzones_bisudo(cur)
    print(f"  -> ktdb_zones_subzone 합계: {n_sudo + n_bisudo}행 삽입")


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"프로젝트 루트: {PROJECT_ROOT}")
    print(f".env 경로:    {env_path}")
    if not env_path.exists():
        print(f"ERROR: .env 파일 없음: {env_path}", file=sys.stderr)
        sys.exit(1)

    # 파일 존재 확인
    for label, path in [
        ("250존 파일", FILE_250),
        ("수도권 소존 파일", FILE_SUDOGWON),
        ("비수도권 소존 파일", FILE_BISUDO),
    ]:
        if not path.exists():
            print(f"ERROR: {label} 없음: {path}", file=sys.stderr)
            sys.exit(1)
        print(f"확인: {label} -> {path.name}")

    print("\nDB 연결 중...")
    try:
        conn = get_connection()
    except Exception as e:
        print(f"ERROR: DB 연결 실패: {e}", file=sys.stderr)
        sys.exit(1)
    print("DB 연결 성공\n")

    try:
        with conn.cursor() as cur:
            seed_zones_17(cur)
            seed_zones_250(cur)
            seed_zones_subzone(cur)
        conn.commit()
        print("\n완료: 모든 존 마스터 데이터 적재 성공")
    except Exception as e:
        conn.rollback()
        print(f"\nERROR: 적재 실패, 롤백 완료: {e}", file=sys.stderr)
        raise
    finally:
        conn.close()

    # 최종 행 수 확인
    print("\n--- 테이블별 최종 행 수 ---")
    try:
        conn2 = get_connection()
        with conn2.cursor() as cur:
            for table in ["ktdb_zones_17", "ktdb_zones_250", "ktdb_zones_subzone"]:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                count = cur.fetchone()[0]
                print(f"  {table}: {count}행")
        conn2.close()
    except Exception as e:
        print(f"행 수 확인 실패: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
