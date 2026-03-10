"""
KTDB 사회경제지표 PostgreSQL 적재 스크립트
- Task 1: 전국지역간 사회경제지표 (250존 시군구 단위) → ktdb_socioeconomic
- Task 2: 권역별 사회경제지표 (부산울산/대구/대전세종충청/광주/제주, 소존 단위) → ktdb_socioeconomic
- Task 3: 수도권 사회경제지표 (TXT, CP949, 소존 단위) → ktdb_socioeconomic

파일 구조 분석:
  전국 250존 xlsx: 6시트 (총인구/5-24세인구/15세이상인구/취업자수/학생인구수/종사자수)
    - 컬럼: 시도코드 | 시군구명 | 시군구일련번호 | 2023년 | 2025년 | ... | 2050년 (wide)
    - zone_level='sigungu', zone_code=시군구일련번호(str)

  권역별 xlsx: 7시트 (총인구/5-24세인구/15세이상인구/취업자수/종사자수/3차산업종사자수/학생인구수)
    - 컬럼: 시도코드 | 읍면동명 | 읍면동일련번호 | 2023년 | ... | 2050년 (wide)
    - zone_level='subzone', zone_code=읍면동일련번호(str)

  수도권 TXT (CP949, space-delimited):
    - 인구수/SUB_POP{yy}.TXT: 일련번호 ZONE_ID 남0~100세 여0~100세 (204 cols → 합계)
    - 종사자수/WORK_POP{yy}.TXT: 일련번호 ZONE_ID col3 col4 (2 값 → 합계)
    - 취업자수/EMP_POP_{yy}.TXT: 일련번호 ZONE_ID col3..col12 (연령별 → 합계)
    - 학생수/STU_POP{yy}.TXT: 일련번호 ZONE_ID 초 중 고 특수 대 (5 값 → 합계)
    - zone_level='subzone', zone_code=ZONE_ID(str)

테이블 스키마:
  ktdb_socioeconomic(id, year, zone_level, zone_code, indicator_type, value, source_file, created_at)
  UNIQUE(year, zone_code, indicator_type)
"""

from __future__ import annotations

import io
import os
import re
import sys
import time
from pathlib import Path
from typing import Iterator

import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Project root & env
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")

import psycopg  # noqa: E402

# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------

def get_conn() -> psycopg.Connection:
    conninfo = (
        f"host={os.environ['DB_HOST']} "
        f"port={os.environ.get('DB_PORT', '5432')} "
        f"dbname={os.environ['DB_NAME']} "
        f"user={os.environ['DB_USER']} "
        f"password={os.environ['DB_PASSWORD']} "
        f"sslmode=require"
    )
    return psycopg.connect(conninfo, prepare_threshold=None)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

YEARS = [2023, 2025, 2030, 2035, 2040, 2045, 2050]
YEAR_COLS = [3, 4, 5, 6, 7, 8, 9]  # 0-indexed column indices for year values in xlsx


def _float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def elapsed(start: float) -> str:
    s = time.time() - start
    return f"{s:.1f}s"


def create_table(conn: psycopg.Connection) -> None:
    """Create ktdb_socioeconomic table if not exists."""
    ddl = """
    CREATE TABLE IF NOT EXISTS ktdb_socioeconomic (
        id            BIGSERIAL PRIMARY KEY,
        year          SMALLINT NOT NULL,
        zone_level    VARCHAR(20) NOT NULL,
        zone_code     VARCHAR(20) NOT NULL,
        indicator_type VARCHAR(30) NOT NULL,
        value         DOUBLE PRECISION,
        source_file   TEXT,
        created_at    TIMESTAMPTZ DEFAULT NOW(),
        UNIQUE (year, zone_code, indicator_type)
    );
    CREATE INDEX IF NOT EXISTS idx_socioeconomic_year
        ON ktdb_socioeconomic (year);
    CREATE INDEX IF NOT EXISTS idx_socioeconomic_zone_code
        ON ktdb_socioeconomic (zone_code);
    CREATE INDEX IF NOT EXISTS idx_socioeconomic_indicator
        ON ktdb_socioeconomic (indicator_type);
    """
    with conn.cursor() as cur:
        cur.execute(ddl)
    conn.commit()
    print("  테이블 ktdb_socioeconomic 생성/확인 완료")


def _already_loaded(cur: psycopg.Cursor, year: int, indicator_type: str, zone_level: str) -> bool:
    """Check if data for a year+indicator+zone_level combo already exists."""
    cur.execute(
        "SELECT 1 FROM ktdb_socioeconomic "
        "WHERE year=%s AND indicator_type=%s AND zone_level=%s LIMIT 1",
        (year, indicator_type, zone_level),
    )
    return cur.fetchone() is not None


COPY_SQL = (
    "COPY ktdb_socioeconomic "
    "(year, zone_level, zone_code, indicator_type, value, source_file) "
    "FROM STDIN"
)


# ---------------------------------------------------------------------------
# Sheet → indicator_type mapping
# Sheet names are garbled in terminal but positional mapping is reliable
# Confirmed order from analysis:
#   전국 250존 (6 sheets): 총인구, 5-24세인구, 15세이상인구, 취업자수, 학생인구수, 종사자수
#   권역별    (7 sheets): 총인구, 5-24세인구, 15세이상인구, 취업자수, 종사자수, 3차산업종사자수, 학생인구수
# ---------------------------------------------------------------------------

NATIONAL_250_SHEET_INDICATORS = [
    "population",           # 총인구
    "population_5_24",      # 5-24세 인구
    "population_15plus",    # 15세 이상 인구
    "employment",           # 취업자수
    "student",              # 학생인구수
    "worker",               # 종사자수
]

REGIONAL_SHEET_INDICATORS = [
    "population",           # 총인구
    "population_5_24",      # 5-24세 인구  (부산/대구/광주: 5-24세; 대전/제주: 5-29세)
    "population_15plus",    # 15세 이상 인구
    "employment",           # 취업자수
    "worker",               # 종사자수
    "worker_tertiary",      # 3차산업 종사자수
    "student",              # 학생인구수
]


# ---------------------------------------------------------------------------
# TASK 1: 전국 250존 사회경제지표
# ---------------------------------------------------------------------------

NATIONAL_250_FILE = (
    PROJECT_ROOT
    / "data/extracted"
    / "2024-OD-PSN-MOD-10 전국지역간 주수단 OD(250존)(2023-2050)"
    / "3. 사회경제지표"
    / "전국지역간_사회경제지표.xlsx"
)


def _rows_national_250(
    ws, year: int, year_col_idx: int, indicator_type: str, zone_level: str, source_file: str
) -> Iterator[tuple]:
    """
    Yield rows from one sheet (wide format: one column per year).
    Columns: 시도코드(0) | 시군구명(1) | 시군구일련번호(2) | 2023년(3) | 2025년(4) | ...
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        zone_code_raw = _int(row[2])
        if zone_code_raw is None:
            continue
        zone_code = str(zone_code_raw)
        value = _float(row[year_col_idx])
        if value is None:
            continue
        yield (year, zone_level, zone_code, indicator_type, value, source_file)


def load_national_250(conn: psycopg.Connection) -> int:
    """Load Task 1: 전국 250존 사회경제지표."""
    print("\n=== TASK 1: 전국 250존 사회경제지표 ===")
    total_rows = 0
    t0 = time.time()

    if not NATIONAL_250_FILE.exists():
        print(f"  ERROR: 파일 없음 → {NATIONAL_250_FILE}")
        return 0

    print(f"  파일: {NATIONAL_250_FILE.name}")
    zone_level = "sigungu"

    wb = openpyxl.load_workbook(NATIONAL_250_FILE, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  시트 수: {len(sheet_names)} → {sheet_names}")

    with conn.cursor() as cur:
        for sheet_idx, indicator_type in enumerate(NATIONAL_250_SHEET_INDICATORS):
            if sheet_idx >= len(sheet_names):
                print(f"  SKIP: sheet index {sheet_idx} 없음 (indicator={indicator_type})")
                continue

            ws = wb[sheet_names[sheet_idx]]

            for year_pos, year in enumerate(YEARS):
                year_col_idx = YEAR_COLS[year_pos]

                if _already_loaded(cur, year, indicator_type, zone_level):
                    print(f"  SKIP: year={year} indicator={indicator_type} already loaded")
                    continue

                t1 = time.time()
                rows = list(_rows_national_250(
                    ws, year, year_col_idx, indicator_type, zone_level,
                    NATIONAL_250_FILE.name
                ))
                n = len(rows)
                if n == 0:
                    print(f"  SKIP: no data — year={year} indicator={indicator_type}")
                    continue

                with cur.copy(COPY_SQL) as copy:
                    for row in rows:
                        copy.write_row(row)
                conn.commit()
                total_rows += n
                print(f"  [{sheet_names[sheet_idx]}] year={year} indicator={indicator_type}: {n:,} rows in {elapsed(t1)}")

    wb.close()
    print(f"\n  TASK 1 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# TASK 2: 권역별 사회경제지표 (xlsx)
# ---------------------------------------------------------------------------

REGIONAL_FILES = [
    (
        "부산울산권",
        PROJECT_ROOT / "data/extracted"
        / "2024-OD-PSN-MOD-12 부산울산권 주수단 OD(2023-2050)"
        / "2024-OD-PSN-MOD-12 부산울산권 주수단 OD(2023-2050)"
        / "3. 사회경제지표"
        / "부산울산권_사회경제지표.xlsx",
    ),
    (
        "대구광역권",
        PROJECT_ROOT / "data/extracted"
        / "2024-OD-PSN-MOD-13 대구광역권 주수단 OD(2023-2050)"
        / "2024-OD-PSN-MOD-13 대구광역권 주수단 OD(2023-2050)"
        / "3. 사회경제지표"
        / "대구광역권_사회경제지표.xlsx",
    ),
    (
        "대전세종충청권",
        PROJECT_ROOT / "data/extracted"
        / "2024-OD-PSN-MOD-14 대전세종충청권 주수단 OD(2023-2050)"
        / "2024-OD-PSN-MOD-14 대전세종충청권 주수단 OD(2023-2050)"
        / "3. 사회경제지표"
        / "대전세종충청권_사회경제지표.xlsx",
    ),
    (
        "광주광역권",
        PROJECT_ROOT / "data/extracted"
        / "2024-OD-PSN-MOD-15 광주광역권 주수단 OD(2023-2050)"
        / "2024-OD-PSN-MOD-15 광주광역권 주수단 OD(2023-2050)"
        / "3. 사회경제지표"
        / "광주광역권_사회경제지표.xlsx",
    ),
    (
        "제주권",
        PROJECT_ROOT / "data/extracted"
        / "2024-OD-PSN-MOD-16 제주권 주수단OD(2023-2050)"
        / "2024-OD-PSN-MOD-16 제주권 주수단OD(2023-2050)"
        / "3. 사회경제지표"
        / "제주권_사회경제지표.xlsx",
    ),
]


def _rows_regional(
    ws, year: int, year_col_idx: int, indicator_type: str, zone_level: str,
    region_name: str, source_file: str
) -> Iterator[tuple]:
    """
    Yield rows from one regional sheet (wide format).
    Columns: 시도코드(0) | 읍면동명(1) | 읍면동일련번호(2) | 2023년(3) | ...
    zone_code = region_prefix + '_' + 일련번호 to avoid collisions across regions
    """
    # Region prefix map (short codes to distinguish same serial numbers across regions)
    prefix_map = {
        "부산울산권": "BU",
        "대구광역권": "DG",
        "대전세종충청권": "DJ",
        "광주광역권": "GJ",
        "제주권": "JJ",
    }
    prefix = prefix_map.get(region_name, region_name[:2])

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        serial_raw = _int(row[2])
        if serial_raw is None:
            continue
        # Encode ZONE_ID as region_prefix + serial to avoid cross-region collisions
        zone_code = f"{prefix}_{serial_raw:05d}"
        value = _float(row[year_col_idx])
        if value is None:
            continue
        yield (year, zone_level, zone_code, indicator_type, value, source_file)


def load_regional(conn: psycopg.Connection) -> int:
    """Load Task 2: 권역별 사회경제지표 (xlsx)."""
    print("\n=== TASK 2: 권역별 사회경제지표 (xlsx) ===")
    total_rows = 0
    t0 = time.time()
    zone_level = "subzone"

    with conn.cursor() as cur:
        for region_name, xlsx_path in REGIONAL_FILES:
            print(f"\n  [{region_name}] {xlsx_path.name}")
            if not xlsx_path.exists():
                print(f"    ERROR: 파일 없음 → {xlsx_path}")
                continue

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            print(f"    시트 수: {len(sheet_names)}")

            for sheet_idx, indicator_type in enumerate(REGIONAL_SHEET_INDICATORS):
                if sheet_idx >= len(sheet_names):
                    print(f"    SKIP: sheet index {sheet_idx} 없음 (indicator={indicator_type})")
                    continue

                ws = wb[sheet_names[sheet_idx]]

                # Use region-specific indicator key to avoid cross-region SKIP
                region_indicator = f"{region_name}_{indicator_type}"

                for year_pos, year in enumerate(YEARS):
                    year_col_idx = YEAR_COLS[year_pos]

                    if _already_loaded(cur, year, region_indicator, zone_level):
                        print(f"    SKIP: year={year} {region_indicator} already loaded")
                        continue

                    t1 = time.time()
                    rows = list(_rows_regional(
                        ws, year, year_col_idx, indicator_type, zone_level,
                        region_name, xlsx_path.name
                    ))
                    n = len(rows)
                    if n == 0:
                        print(f"    SKIP: no data — year={year} {indicator_type}")
                        continue

                    with cur.copy(COPY_SQL) as copy:
                        for row in rows:
                            copy.write_row(row)
                    conn.commit()
                    total_rows += n
                    print(f"    year={year} {indicator_type}: {n:,} rows in {elapsed(t1)}")

            wb.close()

    print(f"\n  TASK 2 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# TASK 3: 수도권 사회경제지표 (TXT, CP949)
# ---------------------------------------------------------------------------

METRO_BASE = (
    PROJECT_ROOT
    / "data/extracted"
    / "2024-OD-PSN-MOD-11 수도권 주수단 OD(2023-2050)"
    / "2024-OD-PSN-MOD-11 수도권 주수단 OD(2023-2050)"
    / "3. 사회경제지표"
)

# year suffix → year value
METRO_YEAR_MAP = {
    "23": 2023,
    "25": 2025,
    "30": 2030,
    "35": 2035,
    "40": 2040,
    "45": 2045,
    "50": 2050,
}

# TXT type definitions:
# (folder_name, file_pattern, indicator_type, value_strategy)
# value_strategy:
#   'sum_all'   → sum all value columns (skip col0=serial, col1=ZONE_ID)
#   'sum_cols'  → sum specific column range
METRO_TXT_TYPES = [
    # folder        file_glob_prefix  indicator_type   value_cols_start  value_cols_end (exclusive, None=all)
    ("인구수",      "SUB_POP",        "population",     2, None),
    ("종사자수",    "WORK_POP",       "worker",         2, None),
    ("취업자수",    "EMP_POP_",       "employment",     2, None),
    ("학생수",      "STU_POP",        "student",        2, None),
]


def _parse_metro_txt_file(txt_path: Path, indicator_type: str,
                          val_start: int, val_end: int | None,
                          year: int, zone_level: str, source_file: str) -> list[tuple]:
    """
    Parse one 수도권 TXT file.
    Row layout (space-delimited): serial ZONE_ID val1 val2 ... valN
    Returns list of (year, zone_level, zone_code, indicator_type, value, source_file)
    """
    rows = []
    with open(txt_path, "r", encoding="cp949", errors="replace") as f:
        lines = f.readlines()

    for line in lines[1:]:  # skip header line
        parts = line.split()
        if len(parts) < 3:
            continue
        try:
            zone_id = str(int(parts[1]))  # ZONE_ID as string
        except (ValueError, IndexError):
            continue

        try:
            vals = parts[val_start:val_end]
            # Strip trailing '.' from Fortran-style floats like '5136.'
            numeric_vals = []
            for v in vals:
                v_clean = v.rstrip('.')
                if v_clean:
                    try:
                        numeric_vals.append(float(v_clean))
                    except ValueError:
                        pass
            if not numeric_vals:
                continue
            total_value = sum(numeric_vals)
        except (ValueError, IndexError):
            continue

        rows.append((year, zone_level, zone_code := f"SU_{zone_id}", indicator_type, total_value, source_file))

    return rows


def load_metro_txt(conn: psycopg.Connection) -> int:
    """Load Task 3: 수도권 사회경제지표 (TXT)."""
    print("\n=== TASK 3: 수도권 사회경제지표 (TXT) ===")
    total_rows = 0
    t0 = time.time()
    zone_level = "subzone"
    region_name = "수도권"

    with conn.cursor() as cur:
        for folder_name, file_prefix, indicator_type, val_start, val_end in METRO_TXT_TYPES:
            folder_path = METRO_BASE / folder_name
            if not folder_path.exists():
                print(f"  ERROR: 폴더 없음 → {folder_path}")
                continue

            print(f"\n  [{folder_name}] indicator={indicator_type}")
            files = sorted(folder_path.iterdir())

            for txt_file in files:
                if not txt_file.suffix.upper() == ".TXT":
                    continue

                # Extract year suffix from filename: SUB_POP23.TXT → '23', EMP_POP_23.TXT → '23'
                name_upper = txt_file.stem.upper()
                # Remove prefix and underscores to get year digits
                year_suffix = re.sub(r'^[A-Z_]+', '', name_upper)
                year = METRO_YEAR_MAP.get(year_suffix)
                if year is None:
                    print(f"    SKIP: 연도 파싱 실패 → {txt_file.name}")
                    continue

                region_indicator = f"{region_name}_{indicator_type}"
                if _already_loaded(cur, year, region_indicator, zone_level):
                    print(f"    SKIP: year={year} {region_indicator} already loaded")
                    continue

                t1 = time.time()
                rows = _parse_metro_txt_file(
                    txt_file, indicator_type, val_start, val_end,
                    year, zone_level, txt_file.name
                )
                n = len(rows)
                if n == 0:
                    print(f"    SKIP: no data — {txt_file.name}")
                    continue

                with cur.copy(COPY_SQL) as copy:
                    for row in rows:
                        copy.write_row(row)
                conn.commit()
                total_rows += n
                print(f"    {txt_file.name} year={year}: {n:,} rows in {elapsed(t1)}")

    print(f"\n  TASK 3 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("KTDB 사회경제지표 PostgreSQL 적재 시작")
    print(f"PROJECT_ROOT: {PROJECT_ROOT}")
    t_start = time.time()

    conn = get_conn()
    print("DB 연결 성공")

    try:
        print("\n테이블 생성/확인 중...")
        create_table(conn)

        r1 = load_national_250(conn)
        r2 = load_regional(conn)
        r3 = load_metro_txt(conn)

    finally:
        conn.close()

    print("\n" + "=" * 60)
    print("전체 완료")
    print(f"  Task 1 (전국 250존):   {r1:>10,} rows")
    print(f"  Task 2 (권역별 xlsx):  {r2:>10,} rows")
    print(f"  Task 3 (수도권 TXT):   {r3:>10,} rows")
    print(f"  합계:                  {r1+r2+r3:>10,} rows")
    print(f"  총 소요시간: {elapsed(t_start)}")


if __name__ == "__main__":
    main()
