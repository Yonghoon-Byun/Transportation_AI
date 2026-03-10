"""Load Task 3: 화물 OD only"""

from __future__ import annotations

import os
import time
from pathlib import Path
from typing import Iterator

import openpyxl
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")

import psycopg


def get_conn() -> psycopg.Connection:
    conninfo = (
        f"host={os.environ['DB_HOST']} "
        f"port={os.environ.get('DB_PORT', '5432')} "
        f"dbname={os.environ['DB_NAME']} "
        f"user={os.environ['DB_USER']} "
        f"password={os.environ['DB_PASSWORD']} "
        f"sslmode=require"
    )
    return psycopg.connect(conninfo, prepare_threshold=None)


def _float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def elapsed(start: float) -> str:
    s = time.time() - start
    return f"{s:.1f}s"


FREIGHT_BASE_FILE = PROJECT_ROOT / "data/extracted/2024-OD-FRE-CAR-00 전국지역간 톤급별 화물자동차 통행량 OD(2023년 기준)/배포용 (기준년도 2023년) 화물자동차OD_2025.11.28.xlsx"
FREIGHT_FUTURE_FILE = PROJECT_ROOT / "data/extracted/2024-OD-FRE-CAR-00 전국지역간 톤급별 화물자동차 통행량 OD(2023년 기준)/배포용 장래년도 화물자동차OD_2025.11.28.xlsx"

FREIGHT_FUTURE_YEARS = [2025, 2030, 2035, 2040, 2045, 2050]


def _already_loaded_freight(cur: psycopg.Cursor, year: int) -> bool:
    cur.execute(
        "SELECT 1 FROM ktdb_od_freight WHERE year=%s LIMIT 1",
        (year,),
    )
    return cur.fetchone() is not None


def _rows_freight(ws, year: int, source_file: str, skip_rows: int = 2) -> Iterator[tuple]:
    """Yield COPY rows from one freight sheet."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        o250 = _int(row[0])
        o17 = _int(row[1])
        d250 = _int(row[2])
        d17 = _int(row[3])
        if o250 is None or d250 is None:
            continue
        small = _float(row[4]) or 0.0
        medium = _float(row[5]) or 0.0
        large = _float(row[6]) or 0.0
        total = _float(row[7]) or 0.0
        yield (year, o250, o17, d250, d17, small, medium, large, total, source_file)


def load_freight_od(conn: psycopg.Connection) -> int:
    """Load Task 3: 화물 OD."""
    print("\n=== TASK 3: 화물 OD -> ktdb_od_freight ===")
    total_rows = 0
    t0 = time.time()

    COPY_SQL = (
        "COPY ktdb_od_freight "
        "(year, origin_zone_250, origin_zone_17, destination_zone_250, destination_zone_17, "
        "tonnage_small, tonnage_medium, tonnage_large, tonnage_total, source_file) "
        "FROM STDIN"
    )

    with conn.cursor() as cur:
        # Base year 2023
        if not FREIGHT_BASE_FILE.exists():
            print(f"  ERROR: 기준년도 파일 없음 -> {FREIGHT_BASE_FILE}")
        else:
            print(f"\n  기준년도 파일: {FREIGHT_BASE_FILE.name}")
            if _already_loaded_freight(cur, 2023):
                print("  SKIP: year=2023 already loaded")
            else:
                wb = openpyxl.load_workbook(FREIGHT_BASE_FILE, read_only=True, data_only=True)
                ws = wb.worksheets[0]
                t1 = time.time()
                rows = list(_rows_freight(ws, 2023, FREIGHT_BASE_FILE.name, skip_rows=2))
                n = len(rows)
                if n > 0:
                    with cur.copy(COPY_SQL) as copy:
                        for row in rows:
                            copy.write_row(row)
                    conn.commit()
                    total_rows += n
                    print(f"  year=2023: {n:,} rows in {elapsed(t1)}")
                else:
                    print("  SKIP: no data rows for year=2023")
                wb.close()

        # Future years
        if not FREIGHT_FUTURE_FILE.exists():
            print(f"  ERROR: 장래년도 파일 없음 -> {FREIGHT_FUTURE_FILE}")
        else:
            print(f"\n  장래년도 파일: {FREIGHT_FUTURE_FILE.name}")
            wb = openpyxl.load_workbook(FREIGHT_FUTURE_FILE, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            print(f"  시트 목록: {sheet_names}")

            for year in FREIGHT_FUTURE_YEARS:
                year_str = f"{year}년"
                ws = None
                if year_str in sheet_names:
                    ws = wb[year_str]
                else:
                    idx = FREIGHT_FUTURE_YEARS.index(year)
                    if idx < len(sheet_names):
                        ws = wb[sheet_names[idx]]
                        print(f"  Using sheet '{sheet_names[idx]}' for year={year}")

                if ws is None:
                    print(f"  SKIP: sheet for year={year} not found")
                    continue

                if _already_loaded_freight(cur, year):
                    print(f"  SKIP: year={year} already loaded")
                    continue

                t1 = time.time()
                rows = list(_rows_freight(ws, year, FREIGHT_FUTURE_FILE.name, skip_rows=2))
                n = len(rows)
                if n == 0:
                    print(f"  SKIP: no data rows for year={year}")
                    continue

                with cur.copy(COPY_SQL) as copy:
                    for row in rows:
                        copy.write_row(row)

                conn.commit()
                total_rows += n
                print(f"  year={year}: {n:,} rows in {elapsed(t1)}")

            wb.close()

    print(f"\n  TASK 3 완료: 총 {total_rows:,} rows in {elapsed(t0)}")
    return total_rows


def main():
    print("TASK 3 화물 OD 적재 시작")
    print(f"PROJECT_ROOT: {PROJECT_ROOT}")
    t_start = time.time()

    conn = get_conn()
    print("DB 연결 성공")

    try:
        r3 = load_freight_od(conn)
    finally:
        conn.close()

    print("\n" + "=" * 60)
    print("TASK 3 완료")
    print(f"  화물 OD: {r3:>10,} rows")
    print(f"  총 소요시간: {elapsed(t_start)}")


if __name__ == "__main__":
    main()
